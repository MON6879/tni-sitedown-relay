import os
import re
import sys
import asyncio
import requests
import openpyxl
from datetime import datetime, timezone, timedelta
from telethon import TelegramClient
from telethon.sessions import StringSession
from dotenv import load_dotenv

# Load local environment
load_dotenv()

API_ID = int(os.environ.get("TELEGRAM_API_ID", "0"))
API_HASH = os.environ.get("TELEGRAM_API_HASH", "")
SESSION_STRING = os.environ.get("TELEGRAM_SESSION", "")
REFUEL_APPS_SCRIPT_URL = os.environ.get("REFUEL_APPS_SCRIPT_URL", "").strip()
REFUEL_CHAT_ID = int(os.environ.get("REFUEL_CHAT_ID", "-5469544739"))

SPREADSHEET_ID = "1JxrA4pJo92Xx_SpwLnOQxphVYwE2iFhLrCOHmyVVuuM"
XLSX_DOWNLOAD_URL = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
XLSX_FILE_PATH = "scratch/sheet_refuel_import.xlsx"

TZ_MM = timezone(timedelta(hours=6, minutes=30))  # Myanmar timezone

def download_spreadsheet():
    print("📥 Downloading spreadsheet to check existing records...")
    os.makedirs("scratch", exist_ok=True)
    r = requests.get(XLSX_DOWNLOAD_URL, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
    r.raise_for_status()
    with open(XLSX_FILE_PATH, "wb") as f:
        f.write(r.content)
    print("✅ Download successful!")

def parse_date_str(val) -> str:
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    s = str(val).strip()
    if len(s) == 10 and s[2] == "/" and s[5] == "/":
        return s
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            dt = datetime.strptime(s[:len(fmt)], fmt)
            return dt.strftime("%d/%m/%Y")
        except ValueError:
            pass
    return s

def load_existing_records():
    existing_plans = set()
    existing_requests = set()
    existing_refueled = set()
    
    if os.path.exists(XLSX_FILE_PATH):
        wb = openpyxl.load_workbook(XLSX_FILE_PATH, data_only=True)
        
        if "Plan refuel" in wb.sheetnames:
            ws = wb["Plan refuel"]
            for r in range(2, ws.max_row + 1):
                date_val = parse_date_str(ws.cell(row=r, column=2).value)
                site = str(ws.cell(row=r, column=4).value or "").strip().upper()
                sender_id = str(ws.cell(row=r, column=8).value or "").strip()
                if sender_id.endswith(".0"):
                    sender_id = sender_id[:-2]
                if date_val and site:
                    existing_plans.add((date_val, site, sender_id))
                    
        if "Team request" in wb.sheetnames:
            ws = wb["Team request"]
            for r in range(2, ws.max_row + 1):
                date_val = parse_date_str(ws.cell(row=r, column=2).value)
                site = str(ws.cell(row=r, column=5).value or "").strip().upper()
                sender_id = str(ws.cell(row=r, column=9).value or "").strip()
                if sender_id.endswith(".0"):
                    sender_id = sender_id[:-2]
                if date_val and site:
                    existing_requests.add((date_val, site, sender_id))
                    
        if "Refueled" in wb.sheetnames:
            ws = wb["Refueled"]
            for r in range(2, ws.max_row + 1):
                date_val = parse_date_str(ws.cell(row=r, column=2).value)  # B
                site = str(ws.cell(row=r, column=4).value or "").strip().upper()  # D
                sender_id = str(ws.cell(row=r, column=19).value or "").strip()  # S
                if sender_id.endswith(".0"):
                    sender_id = sender_id[:-2]
                if date_val and site:
                    existing_refueled.add((date_val, site, sender_id))
                    
    print(f"📊 Existing records loaded: Plans={len(existing_plans)}, Requests={len(existing_requests)}, Refueled={len(existing_refueled)}")
    return existing_plans, existing_requests, existing_refueled

def classify(text: str) -> str | None:
    t = text.lower()
    if "dg type" in t:
        return "REFUELED"
    if "letter" in t and ("submit" in t or "submitted" in t):
        return "LETTER_SUBMIT"
    if "letter" in t and "approved" in t:
        return "LETTER_APPROVED"
    if "plan" in t:
        return "PLAN"
    if "request" in t:
        return "REQUEST"
    return None

def parse_sites_and_qty(text, is_plan):
    results = []
    matched_sites = set()

    pat1 = re.compile(r'TNI(\d{4}(?:_\d+)?)(?:\([^)]*\))?[\s:,+]+(\d+)\s*[Ll]?\b(?!\s*\/)', re.IGNORECASE)
    for m in pat1.finditer(text):
        site_code = "TNI" + m.group(1)
        qty = int(m.group(2))
        results.append({"site": site_code, "qty": qty})
        matched_sites.add(site_code.upper())

    if is_plan:
        pat2 = re.compile(r'TNI(\d{4}(?:_\d+)?)', re.IGNORECASE)
        for m in pat2.finditer(text):
            site_code = "TNI" + m.group(1)
            if site_code.upper() not in matched_sites:
                results.append({"site": site_code, "qty": 440})
                matched_sites.add(site_code.upper())

    seen = {}
    for r in results:
        s = r["site"].upper()
        if s not in seen or r["qty"] > seen[s]:
            seen[s] = r["qty"]

    return [{"site": s, "qty": seen[s]} for s in sorted(seen.keys())]

def parse_refueled_site_and_date(text, msg_date):
    m_site = re.search(r'(?:DG\s*ID|site\s*ID)\s+([^\r\n]+)', text, re.IGNORECASE)
    site_id = ""
    if m_site:
        sm = re.search(r'TNI\d{4}(?:_\d+)?', m_site.group(1), re.IGNORECASE)
        if sm:
            site_id = sm.group(0).upper()
            
    m_date = re.search(r'Date\s*[=:]\s*(\d{1,2}/\d{1,2}/\d{4})', text, re.IGNORECASE)
    date_val = m_date.group(1) if m_date else msg_date.strftime("%d/%m/%Y")
    
    return site_id, date_val

def post_gas(payload):
    if not REFUEL_APPS_SCRIPT_URL:
        print("❌ REFUEL_APPS_SCRIPT_URL not configured")
        return None
    try:
        r = requests.post(REFUEL_APPS_SCRIPT_URL, json=payload, timeout=20)
        return r.json()
    except Exception as e:
        print(f"❌ GAS POST error: {e}")
        return None

async def main():
    if not REFUEL_APPS_SCRIPT_URL:
        print("❌ REFUEL_APPS_SCRIPT_URL not set in env.")
        sys.exit(1)
        
    download_spreadsheet()
    existing_plans, existing_requests, existing_refueled = load_existing_records()
    
    print("📡 Connecting to Telegram Client...")
    client = TelegramClient(StringSession(SESSION_STRING), API_ID, API_HASH)
    
    async with client:
        me = await client.get_me()
        print(f"🔑 Logged in as: @{me.username} ({me.first_name})")
        
        # Get chat entity
        print(f"📡 Fetching chat entity for {REFUEL_CHAT_ID}...")
        entity = await client.get_input_entity(REFUEL_CHAT_ID)
        
        # Fetch last 300 messages
        print("📡 Fetching last 300 messages from history...")
        messages = await client.get_messages(entity, limit=300)
        print(f"Found {len(messages)} messages.")
        
        imported_count = 0
        for msg in reversed(messages):
            if not msg.message:
                continue
                
            text = msg.message
            category = classify(text)
            if not category:
                continue
                
            msg_date = msg.date.astimezone(TZ_MM)
            sender_id = str(msg.sender_id or "")
            sender = ""
            if msg.sender:
                sender = f"{getattr(msg.sender, 'first_name', '') or ''} {getattr(msg.sender, 'last_name', '') or ''}".strip()
                
            if category in ("PLAN", "REQUEST"):
                date_match = re.search(r'(\d{1,2}/\d{1,2}/\d{4})', text)
                date_val = date_match.group(1) if date_match else msg_date.strftime("%d/%m/%Y")
                
                # Normalize date format to dd/mm/yyyy
                parts = date_val.split("/")
                if len(parts) == 3:
                    date_val = f"{parts[0].zfill(2)}/{parts[1].zfill(2)}/{parts[2]}"
                
                entries = parse_sites_and_qty(text, is_plan=(category == "PLAN"))
                if not entries:
                    continue
                    
                # Find missing entries
                missing_entries = []
                existing_set = existing_plans if category == "PLAN" else existing_requests
                for e in entries:
                    site = e["site"].upper()
                    if (date_val, site, sender_id) not in existing_set:
                        missing_entries.append(e)
                        
                if not missing_entries:
                    continue
                    
                # Construct message text for missing entries
                post_text = ""
                if len(missing_entries) == len(entries):
                    post_text = text
                else:
                    label = "plan refuel" if category == "PLAN" else "request refuel"
                    post_text = f"Team {label} {date_val}\n" + "\n".join(f"{e['site']}: {e['qty']}L" for e in missing_entries)
                    
                print(f"🚀 Importing [{category}] date={date_val} sender={sender} missing_sites={[e['site'] for e in missing_entries]}")
                
                payload = {
                    "action": "collect_message",
                    "group_id": str(abs(REFUEL_CHAT_ID)),
                    "text": post_text,
                    "sender": sender or "Unknown",
                    "sender_id": sender_id,
                    "date": msg_date.strftime("%d/%m/%Y %H:%M")
                }
                
                res = post_gas(payload)
                if res and res.get("status") == "ok":
                    print(f"✅ Success! GAS def={res.get('def')}")
                    for e in missing_entries:
                        existing_set.add((date_val, e["site"].upper(), sender_id))
                    imported_count += 1
                else:
                    print(f"❌ Failed: {res}")
                    
            elif category == "REFUELED":
                site_id, date_val = parse_refueled_site_and_date(text, msg_date)
                if not site_id:
                    continue
                    
                parts = date_val.split("/")
                if len(parts) == 3:
                    date_val = f"{parts[0].zfill(2)}/{parts[1].zfill(2)}/{parts[2]}"
                    
                if (date_val, site_id.upper(), sender_id) in existing_refueled:
                    continue
                    
                print(f"🚀 Importing [REFUELED] date={date_val} site={site_id} sender={sender}")
                
                payload = {
                    "action": "collect_message",
                    "group_id": str(abs(REFUEL_CHAT_ID)),
                    "text": text,
                    "sender": sender or "Unknown",
                    "sender_id": sender_id,
                    "date": msg_date.strftime("%d/%m/%Y %H:%M")
                }
                
                res = post_gas(payload)
                if res and res.get("status") == "ok":
                    print(f"✅ Success! GAS def={res.get('def')}")
                    existing_refueled.add((date_val, site_id.upper(), sender_id))
                    imported_count += 1
                else:
                    print(f"❌ Failed: {res}")
                    
            elif category in ("LETTER_SUBMIT", "LETTER_APPROVED"):
                print(f"🚀 Importing [{category}] sender={sender}")
                payload = {
                    "action": "collect_message",
                    "group_id": str(abs(REFUEL_CHAT_ID)),
                    "text": text,
                    "sender": sender or "Unknown",
                    "sender_id": sender_id,
                    "date": msg_date.strftime("%d/%m/%Y %H:%M")
                }
                res = post_gas(payload)
                if res and res.get("status") == "ok":
                    print(f"✅ Success! GAS def={res.get('def')}")
                    imported_count += 1
                else:
                    print(f"❌ Failed: {res}")
                    
        print(f"\n🎉 Completed! Total {imported_count} messages imported/reprocessed successfully.")

if __name__ == "__main__":
    asyncio.run(main())
