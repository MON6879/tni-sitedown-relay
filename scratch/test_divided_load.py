import openpyxl
from datetime import datetime, timezone, timedelta

XLSX_FILE_PATH = "scratch/sheet.xlsx"
TZ_MM = timezone(timedelta(hours=6, minutes=30))

def parse_datetime(val):
    if isinstance(val, datetime):
        if val.tzinfo is None:
            return val.replace(tzinfo=TZ_MM)
        return val
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S"):
            try:
                dt = datetime.strptime(val, fmt)
                return dt.replace(tzinfo=TZ_MM)
            except ValueError:
                pass
    return None

def test_load():
    wb = openpyxl.load_workbook(XLSX_FILE_PATH, data_only=True)
    records = []
    
    # 1. Parse Plan refuel sheet
    if "Plan refuel" in wb.sheetnames:
        ws = wb["Plan refuel"]
        print(f"Loading 'Plan refuel' rows: {ws.max_row}")
        for r in range(2, ws.max_row + 1):
            date_val = ws.cell(row=r, column=2).value
            site = ws.cell(row=r, column=4).value
            qty = ws.cell(row=r, column=5).value
            ts = parse_datetime(ws.cell(row=r, column=6).value) or datetime.now(TZ_MM)
            sender = ws.cell(row=r, column=7).value
            sender_id = ws.cell(row=r, column=8).value
            
            if site:
                records.append({
                    "ts": ts,
                    "date": str(date_val).strip() if date_val else "",
                    "cat": "PLAN",
                    "sender": str(sender).strip() if sender else "",
                    "sender_id": str(sender_id).strip() if sender_id else "",
                    "site": str(site).strip() if site else "",
                    "qty": int(qty) if qty is not None else 0
                })
                
    # 2. Parse Team request sheet
    if "Team request" in wb.sheetnames:
        ws = wb["Team request"]
        print(f"Loading 'Team request' rows: {ws.max_row}")
        for r in range(2, ws.max_row + 1):
            date_val = ws.cell(row=r, column=2).value
            site = ws.cell(row=r, column=5).value
            qty = ws.cell(row=r, column=6).value
            ts = parse_datetime(ws.cell(row=r, column=7).value) or datetime.now(TZ_MM)
            sender = ws.cell(row=r, column=8).value
            sender_id = ws.cell(row=r, column=9).value
            
            if site:
                records.append({
                    "ts": ts,
                    "date": str(date_val).strip() if date_val else "",
                    "cat": "REQUEST",
                    "sender": str(sender).strip() if sender else "",
                    "sender_id": str(sender_id).strip() if sender_id else "",
                    "site": str(site).strip() if site else "",
                    "qty": int(qty) if qty is not None else 0
                })
                
    # 3. Parse Refueled sheet
    if "Refueled" in wb.sheetnames:
        ws = wb["Refueled"]
        print(f"Loading 'Refueled' rows: {ws.max_row}")
        for r in range(2, ws.max_row + 1):
            date_val = ws.cell(row=r, column=4).value
            site = ws.cell(row=r, column=6).value
            qty = ws.cell(row=r, column=17).value
            ts = parse_datetime(ws.cell(row=r, column=20).value) or datetime.now(TZ_MM)
            sender = ws.cell(row=r, column=21).value
            sender_id = ws.cell(row=r, column=22).value
            
            if site:
                records.append({
                    "ts": ts,
                    "date": str(date_val).strip() if date_val else "",
                    "cat": "REFUELED",
                    "sender": str(sender).strip() if sender else "",
                    "sender_id": str(sender_id).strip() if sender_id else "",
                    "site": str(site).strip() if site else "",
                    "qty": int(qty) if qty is not None else 0
                })

    print(f"\nTotal loaded records from separate sheets: {len(records)}")

if __name__ == "__main__":
    test_load()
