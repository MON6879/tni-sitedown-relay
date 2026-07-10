import openpyxl

wb = openpyxl.load_workbook("scratch/sheet.xlsx", data_only=True)
ws = wb["Template"]
print(f"Template max_row: {ws.max_row}, max_column: {ws.max_column}")
for r in range(1, 10):
    row_vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
    print(f"Row {r}: {row_vals}")
