import openpyxl

wb = openpyxl.load_workbook("scratch/sheet.xlsx", data_only=True)
if "Refueled" in wb.sheetnames:
    ws = wb["Refueled"]
    print(f"=== Refueled Sheet — {ws.max_column} columns, {ws.max_row} rows ===")
    print("\n-- HEADERS (Row 1) --")
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        col_letter = ""
        n = c
        while n:
            col_letter = chr((n-1)%26+65) + col_letter
            n = (n-1)//26
        print(f"  {col_letter:2s} (col {c:2d}): {val}")

    print("\n-- DATA (Row 2) --")
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=2, column=c).value
        col_letter = ""
        n = c
        while n:
            col_letter = chr((n-1)%26+65) + col_letter
            n = (n-1)//26
        if val is not None:
            print(f"  {col_letter:2s} (col {c:2d}): {val}")
