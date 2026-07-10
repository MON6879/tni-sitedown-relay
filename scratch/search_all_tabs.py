import openpyxl

wb = openpyxl.load_workbook("scratch/sheet.xlsx", data_only=True)
print("Searching for '6859790680' or any large ID in sheets:")

for name in wb.sheetnames:
    ws = wb[name]
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            val = ws.cell(row=r, column=c).value
            if val is not None:
                val_str = str(val).strip()
                if "685979" in val_str or (val_str.isdigit() and len(val_str) > 8):
                    print(f"Found in sheet '{name}' at row {r}, col {c}: '{val}'")
