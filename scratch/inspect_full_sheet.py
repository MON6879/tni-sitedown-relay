import openpyxl

wb = openpyxl.load_workbook("scratch/sheet.xlsx", data_only=True)

for name in wb.sheetnames:
    ws = wb[name]
    max_r = ws.max_row
    print(f"\nTab: {name} (Rows: {max_r})")
    count = 0
    for r in range(1, max_r + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, 15)]
        if any(v is not None for v in vals):
            count += 1
            if count <= 5:
                print(f"  Row {r}: {vals[:10]}")
    print(f"  Total non-empty rows: {count}")
