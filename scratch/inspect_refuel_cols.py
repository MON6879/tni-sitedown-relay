import openpyxl

wb = openpyxl.load_workbook("scratch/sheet.xlsx", data_only=True)
for sheet_name in ["Refueled", "Team request", "Plan refuel"]:
    ws = wb[sheet_name]
    print(f"\nSheet: {sheet_name} (columns count: {ws.max_column})")
    # Get non-empty columns
    header = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    print("Header:", header)
