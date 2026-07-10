import openpyxl

wb = openpyxl.load_workbook("scratch/sheet.xlsx", data_only=True)
if "Template" in wb.sheetnames:
    ws = wb["Template"]
    print("=== Template Sheet - Row 1 (Header) ===")
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=1, column=c).value
        if val:
            print(f"  Col {c} ({chr(64+c)}): {val}")
    print("\n=== Template Sheet - Row 2 (Sample) ===")
    for c in range(1, ws.max_column + 1):
        val = ws.cell(row=2, column=c).value
        if val:
            print(f"  Col {c} ({chr(64+c)}): {val}")
else:
    print("Template sheet not found!")
    print("Available sheets:", wb.sheetnames)
