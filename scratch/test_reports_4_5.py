import openpyxl
from datetime import datetime, timezone, timedelta

def test_reports():
    wb = openpyxl.load_workbook("scratch/sheet.xlsx", data_only=True)
    
    # 1. Read members from "Telegram ID" sheet
    members = []
    not_joined = []
    if "Telegram ID" in wb.sheetnames:
        ws = wb["Telegram ID"]
        print(f"Reading members from 'Telegram ID' tab. Total rows: {ws.max_row}")
        for r in range(2, ws.max_row + 1):
            tg_id = ws.cell(row=r, column=1).value
            name = ws.cell(row=r, column=6).value
            if not name and not tg_id:
                continue
            
            tg_id_str = str(tg_id).strip() if tg_id is not None else ""
            if not tg_id_str or tg_id_str == "None" or tg_id_str == "0" or not tg_id_str.isdigit():
                not_joined.append(name or "Unknown")
            else:
                members.append({"id": tg_id_str, "name": name or "Unknown"})
                
        print(f"Valid members found: {len(members)}")
        print(f"Not joined members found: {len(not_joined)}")
        print("Sample not joined:", not_joined[:5])
        print("Sample members:", members[:5])
    else:
        print("Telegram ID tab not found")
        
    # 2. Read Column G from Template tab
    target_ids = []
    if "Template" in wb.sheetnames:
        ws = wb["Template"]
        print(f"\nReading target IDs from 'Template' tab. Total rows: {ws.max_row}")
        for r in range(3, ws.max_row + 1):  # Starting from row 3
            val = ws.cell(row=r, column=7).value
            if val is not None:
                val_str = str(val).strip()
                if val_str.isdigit() and len(val_str) > 8:
                    target_ids.append(val_str)
        print(f"Target IDs from Template Col G: {target_ids}")
    else:
        print("Template tab not found")
        
    # 3. Read PlanRefuel data
    records = []
    if "PlanRefuel" in wb.sheetnames:
        ws = wb["PlanRefuel"]
        print(f"\nReading 'PlanRefuel' tab. Total rows: {ws.max_row}")
        # Note: since this is a new sheet, let's print if it has data
        for r in range(2, ws.max_row + 1):
            ts = ws.cell(row=r, column=1).value
            cat = ws.cell(row=r, column=3).value
            sender_id = ws.cell(row=r, column=6).value
            site = ws.cell(row=r, column=7).value
            if ts:
                records.append({
                    "ts": ts,
                    "cat": str(cat).strip(),
                    "sender_id": str(sender_id).strip() if sender_id else "",
                    "site": str(site).strip() if site else ""
                })
        print(f"Total plan/request records: {len(records)}")
    else:
        print("PlanRefuel tab not found (since it's a new tab we will create)")

if __name__ == "__main__":
    test_reports()
