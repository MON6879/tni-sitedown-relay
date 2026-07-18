import openpyxl

wb = openpyxl.load_workbook("scratch/sheet_refuel.xlsx", data_only=True)
if "Telegram ID" in wb.sheetnames:
    ws = wb["Telegram ID"]
    print("Telegram ID Sheet Header and Rows:")
    for r in range(1, min(10, ws.max_row + 1)):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
        print(f"Row {r}: {row_vals}")
else:
    print("Telegram ID sheet not found")
