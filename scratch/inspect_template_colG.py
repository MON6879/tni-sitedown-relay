import openpyxl

wb = openpyxl.load_workbook("scratch/sheet.xlsx", data_only=True)
ws = wb["Template"]
print("Values in Column G of Template sheet:")
for r in range(1, ws.max_row + 1):
    val = ws.cell(row=r, column=7).value
    if val is not None:
        print(f"Row {r}: '{val}' (type: {type(val)})")
