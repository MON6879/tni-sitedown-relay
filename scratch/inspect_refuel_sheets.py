import openpyxl

wb = openpyxl.load_workbook("scratch/sheet.xlsx", data_only=True)

for sheet_name in ["Template", "Plan refuel"]:
    if sheet_name in wb.sheetnames:
        print(f"\n--- Sheet: {sheet_name} ---")
        ws = wb[sheet_name]
        for r_idx in range(1, 15):
            row_vals = [ws.cell(row=r_idx, column=c_idx).value for c_idx in range(1, 10)]
            if any(row_vals):
                print(f"Row {r_idx}: {row_vals}")
    else:
        print(f"\nSheet {sheet_name} not found")
