import openpyxl

wb = openpyxl.load_workbook("scratch/sheet.xlsx", data_only=True)
if "PlanRefuel" in wb.sheetnames:
    ws = wb["PlanRefuel"]
    print(f"PlanRefuel sheet found! Total rows: {ws.max_row}")
    for r in range(1, ws.max_row + 1):
        row_vals = [ws.cell(row=r, column=c).value for c in range(1, 10)]
        print(f"Row {r}: {row_vals}")
else:
    print("PlanRefuel sheet NOT found in downloaded sheet.xlsx")
