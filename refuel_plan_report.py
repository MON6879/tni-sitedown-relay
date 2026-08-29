"""
refuel_plan_report.py
======================
Tạo 5 báo cáo so sánh/thống kê Refuel Plan từ Google Sheet và gửi vào group Telegram.

Báo cáo 1: Tần suất gửi Plan của từng trạm (3 ngày / 7 ngày / 1 tháng)
Báo cáo 2: So sánh Plan hôm nay vs Refueled thực tế (chạy lúc 18:00)
Báo cáo 3: So sánh Plan hôm nay vs Team Request (chạy lúc 18:00)
Báo cáo 4: Thống kê Request refuel theo thành viên (từ sheet Telegram ID) + Ai chưa tham gia (chạy lúc 22:00)
Báo cáo 5: Thống kê Plan gửi theo danh sách ID tại cột G và tên tại cột H sheet Template (chạy lúc 22:00)

Cách chạy:
  python refuel_plan_report.py --report 1
  python refuel_plan_report.py --report 2
  python refuel_plan_report.py --report 3
  python refuel_plan_report.py --report 4
  python refuel_plan_report.py --report 5
  python refuel_plan_report.py            # Chạy toàn bộ 5 báo cáo
"""
import os, sys, argparse, requests
import openpyxl
from datetime import datetime, timezone, timedelta
from dotenv import load_dotenv
from tg_utils import tg_send_fresh

load_dotenv()  # Load .env local (GitHub Actions đã có env vars sẵn)

# Cấu hình bot và chat ID mặc định của group 9 TNI REQUEST REFUEL
REFUEL_BOT_TOKEN    = os.getenv("REFUEL_BOT_TOKEN", "8811503647:AAEVIToiaPbDeNTUPLsoI5xhdnufKdChsME")
REFUEL_CHAT_ID      = "-5469544739"   # Group 9 TNI REQUEST REFUEL
SPREADSHEET_ID      = "1JxrA4pJo92Xx_SpwLnOQxphVYwE2iFhLrCOHmyVVuuM"
XLSX_DOWNLOAD_URL   = f"https://docs.google.com/spreadsheets/d/{SPREADSHEET_ID}/export?format=xlsx"
XLSX_FILE_PATH      = "scratch/sheet_refuel.xlsx"

TZ_MM = timezone(timedelta(hours=6, minutes=30))  # Myanmar UTC+6:30


# ── Helper functions ─────────────────────────────────────────────────────────

# GAS URL để lưu message_id (BotState)
REFUEL_GAS_URL = (
    os.getenv("APPS_SCRIPT_URL") or
    os.getenv("REFUEL_APPS_SCRIPT_URL") or ""
).strip()


# Map report_key → title prefix dòng đầu tiên (dùng cho Telethon delete-by-title)
REPORT_TITLE_PREFIX = {
    "report1": "🔄 [Report 1]",
    "report2": "📊 [Report 2]",
    "report5": "📋 [Report 5]",
}

def tg_send(text: str, report_key: str = "") -> bool:
    """Gửi tin nhắn báo cáo lên Telegram group và tự động xóa tin nhắn cũ cùng tiêu đề."""
    prefix = REPORT_TITLE_PREFIX.get(report_key, "")
    state_key = f"refuel_plan_{report_key}_{REFUEL_CHAT_ID}"
    try:
        new_id = tg_send_fresh(
            chat_id=REFUEL_CHAT_ID,
            text=text,
            state_key=state_key,
            parse_mode="HTML",
            title_prefix=prefix,
            bot_token=REFUEL_BOT_TOKEN
        )
        if new_id:
            print(f"✅ Report {report_key} sent fresh (msg_id={new_id}) to {REFUEL_CHAT_ID}")
            return True
        else:
            print(f"❌ Failed to send Report {report_key} fresh", file=sys.stderr)
            return False
    except Exception as e:
        print(f"❌ Telegram send exception: {e}", file=sys.stderr)
        return False


def fmt_row_compare_5(col_team: str, col_site: str, col_b: str, col_c: str, col_d: str) -> str:
    """Format dòng so sánh 5 cột (Team | Site ID | Req/Plan | Plan/Ref | Diff) dạng monospace."""
    return f"<code>{col_team:<6} | {col_site:<10} | {col_b:>5} | {col_c:>5} | {col_d:>5}</code>"



def download_spreadsheet() -> bool:
    """Tải file Excel từ Google Sheets."""
    print("📥 Downloading spreadsheet...")
    try:
        os.makedirs("scratch", exist_ok=True)
        r = requests.get(XLSX_DOWNLOAD_URL, headers={"User-Agent": "Mozilla/5.0"}, timeout=30)
        r.raise_for_status()
        with open(XLSX_FILE_PATH, "wb") as f:
            f.write(r.content)
        print("✅ Download successful!")
        return True
    except Exception as e:
        print(f"❌ Download failed: {e}", file=sys.stderr)
        return False


def parse_datetime(val) -> datetime | None:
    """Đảm bảo định dạng datetime có múi giờ Myanmar."""
    if isinstance(val, datetime):
        if val.tzinfo is None:
            return val.replace(tzinfo=TZ_MM)
        return val
    if isinstance(val, str):
        val = val.strip()
        for fmt in ("%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S"):
            try:
                dt = datetime.strptime(val, fmt)
                return dt.replace(tzinfo=TZ_MM)
            except ValueError:
                pass
    return None


def parse_date_str(val) -> str:
    """Chuyển đổi giá trị ngày (datetime hoặc string) sang chuỗi dd/MM/YYYY."""
    if val is None:
        return ""
    if isinstance(val, datetime):
        return val.strftime("%d/%m/%Y")
    s = str(val).strip()
    if len(s) == 10 and s[2] == "/" and s[5] == "/":
        return s
    for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:len(fmt)], fmt).strftime("%d/%m/%Y")
        except ValueError:
            pass
    return s


def parse_date_to_datetime(val) -> datetime | None:
    """Parse bất kỳ định dạng ngày nào về datetime object (không kèm timezone) để so sánh."""
    if val is None:
        return None
    if isinstance(val, datetime):
        return val
    s = str(val).strip()
    if not s or s.lower() == "none":
        return None
    # Thử các định dạng ngày phổ biến
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d/%m/%Y %H:%M", "%d/%m/%Y %H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%dT%H:%M:%S"):
        try:
            return datetime.strptime(s[:len(fmt)], fmt)
        except ValueError:
            pass
    return None


def safe_int(val) -> int:
    """Chuyển đổi giá trị số nguyên an toàn — trả 0 nếu không parse được."""
    if val is None:
        return 0
    if isinstance(val, (int, float)):
        return int(val)
    try:
        return int(str(val).strip())
    except (ValueError, TypeError):
        return 0


def fmt_row_freq(col_a: str, col_b: str, col_c: str, col_d: str) -> str:
    """Format dòng tần suất (Site/Name | 3D | 7D | 1M) dạng monospace có gạch dọc '|'."""
    return f"<code>{col_a:<12} | {col_b:>5} | {col_c:>5} | {col_d:>6}</code>"


def fmt_row_compare(col_a: str, col_b: str, col_c: str, col_d: str) -> str:
    """Format dòng so sánh lượng dầu (Site | Plan/Req | Filled/Plan | Diff) dạng monospace có gạch dọc '|'."""
    return f"<code>{col_a:<12} | {col_b:>5} | {col_c:>6} | {col_d:>6}</code>"


# ── Load spreadsheet data ───────────────────────────────────────────────────

class RefuelData:
    def __init__(self):
        self.members = []          # list of dict: {id, name} — từ sheet "Telegram ID"
        self.not_joined = []       # list of str (names)
        self.target_members = []   # list of dict: {id, name} — từ Template col G & H
        self.lettel_persons = []   # list of dict: {id, name} — từ Template col J & K
        self.records = []          # list of dict: {ts, date, cat, sender, sender_id, site, qty}
        self.letter_approved = ""  # ngày Government Approved mới nhất (col C)
        self.letter_submitted = "" # ngày Letter Submitted mới nhất (col B)
        self.ft_monitors = []      # list of dict: {date, ft_name, site_id, qty} — từ sheet "FT follow monitor"
 
        if not os.path.exists(XLSX_FILE_PATH):
            if not download_spreadsheet():
                return
 
        try:
            wb = openpyxl.load_workbook(XLSX_FILE_PATH, data_only=True)
            self._parse_members(wb)
            self._parse_targets(wb)
            self._parse_lettel(wb)
            self._parse_lettel_progress(wb)
            self._parse_ft_monitors(wb)
            self._parse_records(wb)
        except Exception as e:
            print(f"❌ Error loading Excel data: {e}", file=sys.stderr)

    def _parse_members(self, wb):
        if "Telegram ID" in wb.sheetnames:
            ws = wb["Telegram ID"]
            for r in range(2, ws.max_row + 1):
                tg_id = ws.cell(row=r, column=1).value   # Cột A: Telegram ID

                # Lấy tên từ cột F (cột 6), tránh lấy số điện thoại ở cột B (cột 2)
                name = None
                val_f = ws.cell(row=r, column=6).value
                if val_f and str(val_f).strip() and str(val_f).strip().lower() not in ("none", "0"):
                    name = str(val_f).strip()
                else:
                    # Fallback sang các cột khác trừ cột B (2)
                    for col in (3, 4, 5):
                        val = ws.cell(row=r, column=col).value
                        if val and str(val).strip() and str(val).strip().lower() not in ("none", "0"):
                            name = str(val).strip()
                            break

                if not name and not tg_id:
                    continue  # Dòng trống hoàn toàn → bỏ qua

                tg_id_str = str(tg_id).strip() if tg_id is not None else ""
                if not tg_id_str or tg_id_str == "None" or tg_id_str == "0" or not tg_id_str.isdigit():
                    self.not_joined.append(name or "Unknown")
                else:
                    self.members.append({"id": tg_id_str, "name": name or "Unknown"})

    def _parse_targets(self, wb):
        if "Template" in wb.sheetnames:
            ws = wb["Template"]
            for r in range(2, ws.max_row + 1):  # Quét từ dòng 2
                val = ws.cell(row=r, column=7).value     # Cột G: ID telegram
                name = ws.cell(row=r, column=8).value    # Cột H: Tên
                if val is not None:
                    val_str = str(val).strip()
                    # Bỏ qua dòng tiêu đề mẫu nếu có chữ "id telegram"
                    if val_str.lower() == "id telegram":
                        continue
                    if val_str.isdigit() and len(val_str) > 8:
                        self.target_members.append({
                            "id": val_str,
                            "name": str(name).strip() if name else f"ID:{val_str[-6:]}"
                        })

    def _parse_lettel(self, wb):
        """Đọc Template col J (Telegram ID) & K (Tên nhân viên phụ trách)."""
        sheet_name = "Template"
        if sheet_name not in wb.sheetnames:
            return
        ws = wb[sheet_name]
        for r in range(2, ws.max_row + 1):
            tg_id = ws.cell(row=r, column=10).value   # J: Telegram ID nhân viên
            name  = ws.cell(row=r, column=11).value   # K: Tên nhân viên
            if not tg_id and not name:
                continue
            tg_id_str = str(tg_id).strip() if tg_id is not None else ""
            name_str  = str(name).strip()  if name  is not None else ""
            if tg_id_str.isdigit() and len(tg_id_str) > 5 and name_str:
                self.lettel_persons.append({"id": tg_id_str, "name": name_str})

    def _parse_lettel_progress(self, wb):
        """Đọc Lettel Progress col B (Date Letter Submit) và col C (Date Leter Approved) — lấy giá trị mới nhất."""
        if "Lettel Progress" not in wb.sheetnames:
            return
        ws = wb["Lettel Progress"]
        latest_submitted = ""
        latest_approved  = ""
        max_sub_dt = None
        max_app_dt = None
        for r in range(2, ws.max_row + 1):
            val_b = ws.cell(row=r, column=2).value  # B: Date Letter Submit
            val_c = ws.cell(row=r, column=3).value  # C: Date Letter Approved
            if val_b:
                dt_b = parse_date_to_datetime(val_b)
                if dt_b:
                    if max_sub_dt is None or dt_b > max_sub_dt:
                        max_sub_dt = dt_b
                        latest_submitted = parse_date_str(val_b)
            if val_c:
                dt_c = parse_date_to_datetime(val_c)
                if dt_c:
                    if max_app_dt is None or dt_c > max_app_dt:
                        max_app_dt = dt_c
                        latest_approved = parse_date_str(val_c)
        self.letter_submitted = latest_submitted
        self.letter_approved  = latest_approved
 
    def _parse_ft_monitors(self, wb):
        """Đọc tab 'FT follow monitor' và lấy danh sách những người đi theo giám sát."""
        self.ft_monitors = []
        if "FT follow monitor" not in wb.sheetnames:
            return
        ws = wb["FT follow monitor"]
        for r in range(2, ws.max_row + 1):
            date_val = ws.cell(row=r, column=2).value  # B: Date
            ft_name  = ws.cell(row=r, column=3).value  # C: FT Name
            site_id  = ws.cell(row=r, column=4).value  # D: Site ID
            qty      = ws.cell(row=r, column=5).value  # E: Refuel Qty
            
            if ft_name:
                date_str = ""
                if isinstance(date_val, datetime):
                    date_str = date_val.strftime("%d/%m/%Y")
                elif date_val:
                    date_str = str(date_val).strip()
                
                self.ft_monitors.append({
                    "date": date_str,
                    "ft_name": str(ft_name).strip(),
                    "site_id": str(site_id).strip() if site_id else "",
                    "qty": str(qty).strip() if qty else ""
                })
 
    def _parse_records(self, wb):
        # 1. Parse Plan refuel sheet
        if "Plan refuel" in wb.sheetnames:
            ws = wb["Plan refuel"]
            for r in range(2, ws.max_row + 1):
                date_val = ws.cell(row=r, column=2).value
                site = ws.cell(row=r, column=4).value
                qty = ws.cell(row=r, column=5).value
                ts = parse_datetime(ws.cell(row=r, column=6).value) or datetime.now(TZ_MM)
                sender = ws.cell(row=r, column=7).value
                sender_id = ws.cell(row=r, column=8).value
                
                if site:
                    team_val = ws.cell(row=r, column=3).value  # col C: Name Team Plan
                    self.records.append({
                        "ts": ts,
                        "date": parse_date_str(date_val),
                        "cat": "PLAN",
                        "team": str(team_val).strip() if team_val else "",
                        "sender": str(sender).strip() if sender else "",
                        "sender_id": str(sender_id).strip() if sender_id else "",
                        "site": str(site).strip() if site else "",
                        "qty": safe_int(qty)
                    })
                    
        # 2. Parse Team request sheet
        if "Team request" in wb.sheetnames:
            ws = wb["Team request"]
            for r in range(2, ws.max_row + 1):
                date_val = ws.cell(row=r, column=2).value
                team_val = ws.cell(row=r, column=4).value  # col D: Name Team
                site = ws.cell(row=r, column=5).value
                qty = ws.cell(row=r, column=6).value
                ts = parse_datetime(ws.cell(row=r, column=7).value) or datetime.now(TZ_MM)
                sender = ws.cell(row=r, column=8).value
                sender_id = ws.cell(row=r, column=9).value
                
                if site:
                    self.records.append({
                        "ts": ts,
                        "date": parse_date_str(date_val),
                        "cat": "REQUEST",
                        "team": str(team_val).strip() if team_val else "",
                        "sender": str(sender).strip() if sender else "",
                        "sender_id": str(sender_id).strip() if sender_id else "",
                        "site": str(site).strip() if site else "",
                        "qty": safe_int(qty)
                    })
                    
        # 3. Parse Refueled sheet
        if "Refueled" in wb.sheetnames:
            ws = wb["Refueled"]
            for r in range(2, ws.max_row + 1):
                date_val = ws.cell(row=r, column=2).value
                team_val = ws.cell(row=r, column=5).value  # col E: Team
                site = ws.cell(row=r, column=4).value
                qty = ws.cell(row=r, column=14).value       # col N: Actual Filled Qty(L)
                ts = parse_datetime(ws.cell(row=r, column=17).value) or datetime.now(TZ_MM)
                sender = ws.cell(row=r, column=18).value
                sender_id = ws.cell(row=r, column=19).value
                
                if site:
                    self.records.append({
                        "ts": ts,
                        "date": parse_date_str(date_val),
                        "cat": "REFUELED",
                        "team": str(team_val).strip() if team_val else "",
                        "sender": str(sender).strip() if sender else "",
                        "sender_id": str(sender_id).strip() if sender_id else "",
                        "site": str(site).strip() if site else "",
                        "qty": safe_int(qty)
                    })


# ── Reports implementation ──────────────────────────────────────────────────

def report_1(data: RefuelData):
    """
    Báo cáo tổng hợp kết hợp:
    1. Letter Progress + FT follow monitor
    2. Bảng đối soát 5 cột phân theo Team kèm tổng số Sites: 🏷️ Team X (N Sites)
    3. Thống kê trạm lệch (Diff sites summary)
    4. Thống kê trạm Team Request chưa có Plan hoặc chưa Refuel (Pending requests)
    """
    print("🔄 Generating Combined Refuel Plan & Progress Daily Report...")
    now = datetime.now(TZ_MM)
    today_str = now.strftime("%d/%m/%Y")

    # ── Letter Progress ──
    submit_line   = f"  📤 The letter was submitted to the Government for approval on: <b>{data.letter_submitted or 'N/A'}</b>"
    approved_line = f"  ✅ The government approved the oil transport letter on: <b>{data.letter_approved or 'N/A'}</b>"

    # ── FT follow monitor ──
    ft_today = []
    for ft in data.ft_monitors:
        ft_date = ft["date"]
        try:
            parts = ft_date.split("/")
            if len(parts) == 3:
                ft_date = f"{int(parts[0]):02d}/{int(parts[1]):02d}/{int(parts[2])}"
        except Exception:
            pass
        today_norm = ""
        try:
            parts_today = today_str.split("/")
            if len(parts_today) == 3:
                today_norm = f"{int(parts_today[0]):02d}/{int(parts_today[1]):02d}/{int(parts_today[2])}"
        except Exception:
            today_norm = today_str
        if ft_date == today_norm:
            ft_today.append(ft["ft_name"])
    ft_names_today = sorted(list(set(ft_today)))
    ft_str = ", ".join(ft_names_today) if ft_names_today else "None"

    # ── Gather data for today by Team ──
    teams_list = ["Team 1", "Team 2", "Team 3", "Team 4"]
    team_map: dict[str, dict[str, dict[str, int]]] = {t: {} for t in teams_list}

    def get_team_key(raw_team: str) -> str:
        s = raw_team.lower()
        if "team 1" in s or "team1" in s: return "Team 1"
        if "team 2" in s or "team2" in s: return "Team 2"
        if "team 3" in s or "team3" in s: return "Team 3"
        if "team 4" in s or "team4" in s: return "Team 4"
        return "Team 1"

    for r in data.records:
        if r["date"] != today_str or not r["site"]:
            continue
        team = get_team_key(r.get("team", ""))
        site = r["site"]
        if site not in team_map[team]:
            team_map[team][site] = {"plan": 0, "refueled": 0, "req": 0}

        if r["cat"] == "PLAN":
            team_map[team][site]["plan"] += r["qty"]
        elif r["cat"] == "REFUELED":
            team_map[team][site]["refueled"] += r["qty"]
        elif r["cat"] == "REQUEST":
            team_map[team][site]["req"] += r["qty"]

    # Header
    lines = [
        f"🔄 <b>[Report 1] PLAN & PROGRESS DAILY REPORT — {today_str}</b>",
        f"⏰ {now.strftime('%H:%M')} Myanmar",
        "",
        "📝 <b>Letter Progress:</b>",
        submit_line,
        approved_line,
        f"  👥 FT follow monitor: <b>{ft_str}</b>",
        "",
        "🟩 Match  🟨 Diff qty  🟥 Not filled  🟦 Extra filled  🟣 Req only",
        fmt_row_compare_5("Team", "Site ID", "Plan", "Refueled", "Diff"),
        "<code>" + "───────┼───────────┼───────┼───────┼──────" + "</code>",
    ]

    green_total = yellow_total = red_total = blue_total = purple_total = 0
    diff_by_team: dict[str, list[str]] = {t: [] for t in teams_list}
    req_only_by_team: dict[str, list[str]] = {t: [] for t in teams_list}
    total_sites_count = 0

    for team in teams_list:
        sites_data = team_map[team]
        if not sites_data:
            continue

        total_sites_count += len(sites_data)
        lines.append(f"\n🏷 <b>{team} ({len(sites_data)} Sites)</b>")
        for site in sorted(sites_data.keys()):
            p = sites_data[site]["plan"]
            fill = sites_data[site]["refueled"]
            q = sites_data[site]["req"]
            diff = p - fill

            if p > 0 and fill > 0 and diff == 0:
                icon = "🟩"; green_total += 1
            elif p > 0 and fill > 0 and diff != 0:
                icon = "🟨"; yellow_total += 1
                diff_by_team[team].append(f"{site} ({p}L vs {fill}L)")
            elif p > 0 and fill == 0:
                icon = "🟥"; red_total += 1
                diff_by_team[team].append(f"{site} (Unfilled {p}L)")
            elif p == 0 and fill > 0:
                icon = "🟦"; blue_total += 1
                diff_by_team[team].append(f"{site} (Unplanned {fill}L)")
            elif p == 0 and fill == 0 and q > 0:
                icon = "🟣"; purple_total += 1
                req_only_by_team[team].append(f"{site} ({q}L)")
            else:
                icon = "⬜"

            diff_str = "=" if diff == 0 else f"{diff:+d}L"
            lines.append(f"{icon} {fmt_row_compare_5(team, site, f'{p}L', f'{fill}L', diff_str)}")

    lines.append("<code>" + "───────┴───────────┴───────┴───────┴──────" + "</code>")

    # Conclusion / Summary of Diff sites per Team
    diff_summary_lines = []
    for team in teams_list:
        diff_list = diff_by_team[team]
        sites_data = team_map[team]
        if diff_list:
            diff_summary_lines.append(f"  • <b>{team}</b> ({len(diff_list)} sites): " + ", ".join(diff_list[:5]))
            if len(diff_list) > 5:
                diff_summary_lines.append(f"    ... +{len(diff_list)-5} more sites")
        elif sites_data:
            diff_summary_lines.append(f"  • <b>{team}</b>: 100% Matched ✅")

    if diff_summary_lines:
        lines.append("\n📌 <b>DIFF SITES SUMMARY BY TEAM:</b>")
        lines.extend(diff_summary_lines)

    # Pending / Unfulfilled Team Requests section
    has_req_pending = any(len(v) > 0 for v in req_only_by_team.values())
    lines.append("\n⚠️ <b>PENDING TEAM REQUESTS (Chưa có Plan / Chưa Refuel):</b>")
    if has_req_pending:
        for team, req_list in req_only_by_team.items():
            if req_list:
                lines.append(f"  • <b>{team}</b> ({len(req_list)} sites): " + ", ".join(req_list[:5]))
                if len(req_list) > 5:
                    lines.append(f"    ... +{len(req_list)-5} more sites")
    else:
        lines.append("  🎉 100% Team requests have been planned / refueled.")

    lines.append(f"\n🟩 <b>{green_total}</b>  🟨 <b>{yellow_total}</b>  🟥 <b>{red_total}</b>  🟦 <b>{blue_total}</b>  🟣 <b>{purple_total}</b>")
    lines.append("\n🤖 <i>Auto report — Refuel Plan System</i>")

    tg_send("\n".join(lines), "report1")
    print("✅ Combined Report 1 sent.")


def report_2(data: RefuelData):
    """Gộp chung vào Report 1."""
    report_1(data)




def report_4(data: RefuelData):
    """Alias cho report_2 để tương thích với các script cũ."""
    report_2(data)


def report_5(data: RefuelData):
    print("📋 Generating Report 5 — Members Not Joined Telegram Group...")
    now = datetime.now(TZ_MM)

    lines = [
        f"📋 <b>[Report 5] MEMBERS NOT YET JOINED</b>",
        f"📅 {now.strftime('%d/%m/%Y %H:%M')} (Myanmar)",
    ]

    not_joined = data.not_joined   # list of str (tên, chưa có Telegram ID)
    joined     = data.members      # list of {id, name} (đã có ID)

    if not not_joined:
        lines += [
            f"\n✅ All <b>{len(joined)}</b> members have joined the Telegram group!",
            "\n🤖 <i>Auto report — Refuel Plan System</i>"
        ]
        tg_send("\n".join(lines), "report5")
        print("✅ Report 5 sent — all joined.")
        return

    lines += [
        f"\n❌ <b>{len(not_joined)}</b> members NOT yet joined | ✅ Joined: <b>{len(joined)}</b>",
        f"<code>{'No':<3} {'Name':<25}</code>",
        "<code>" + "────┬────────────────────────────" + "</code>",
    ]
    for i, name in enumerate(sorted(not_joined), 1):
        lines.append(f"<code>{i:<3} {name[:25]:<25}</code>")

    lines += [
        "<code>" + "────┴────────────────────────────" + "</code>",
        "\n🤖 <i>Auto report — Refuel Plan System</i>"
    ]
    tg_send("\n".join(lines), "report5")
    print(f"✅ Report 5 sent — {len(not_joined)} not joined.")



# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    # Preprocess sys.argv to split '21' into '2' '1' if present
    new_argv = []
    for arg in sys.argv:
        if arg == "21":
            new_argv.extend(["2", "1"])
        else:
            new_argv.append(arg)
    sys.argv = new_argv

    parser = argparse.ArgumentParser(description="TNI Refuel Plan Reports")
    parser.add_argument("--report", type=int, choices=[1, 2, 4, 5, 21], nargs="+",
                        help="Report numbers (1, 2, 3, 4, 5). Omit to run default.")
    args = parser.parse_args()

    # Tải và parse dữ liệu trước khi chạy báo cáo
    download_spreadsheet()
    data = RefuelData()

    raw_reports = args.report if args.report else [1, 2]
    reports_to_run = []
    for r in raw_reports:
        if r == 21:
            reports_to_run.extend([2, 1])
        else:
            reports_to_run.append(r)

    if 1 in reports_to_run:
        report_1(data)
    if 2 in reports_to_run or 4 in reports_to_run:
        report_2(data)

    if 5 in reports_to_run:
        report_5(data)

    print("🎉 All tasks finished successfully.")


if __name__ == "__main__":
    main()
